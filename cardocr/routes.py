"""HTML and JSON routes for the business-card manager."""

from __future__ import annotations

import csv
import io
import re
import time
import uuid
from pathlib import Path
from typing import Any

from flask import (
    Blueprint,
    Response,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
)

from .database import Database
from .gemini_field_classifier import gemini_classifier
from .image_processing import (
    InvalidImageError,
    as_data_url,
    decode_image,
    detect_and_rectify,
    encode_jpeg,
    image_metadata,
    resize_for_ocr,
)
from .ocr_engine import OCRLine, OCRUnavailableError, engine
from .parser import ParsedCard


web = Blueprint("web", __name__)
FRONTEND_DIR = Path(__file__).resolve().parent / "static"
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
EXPORT_HEADERS = {
    "id": "번호",
    "name": "이름",
    "company": "회사",
    "job_title": "직책",
    "phone": "전화번호 1",
    "phone2": "전화번호 2",
    "fax": "팩스",
    "email": "이메일",
    "website": "웹사이트",
    "address": "주소",
    "memo": "메모",
    "created_at": "등록일",
    "updated_at": "수정일",
}


def _database() -> Database:
    return current_app.extensions["database"]


def _error(message: str, status: int = 400, code: str = "BAD_REQUEST"):
    return jsonify({"ok": False, "error": {"code": code, "message": message}}), status


def _payload() -> dict[str, Any]:
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        raise ValueError("JSON 요청 본문이 필요합니다.")
    return data


def _validate_contact(data: dict[str, Any]) -> str | None:
    values = [
        str(data.get(field, "") or "").strip()
        for field in ("name", "company", "phone", "phone2", "fax", "email")
    ]
    if not any(values):
        return "이름, 회사, 전화번호, 이메일 중 하나 이상을 입력해 주세요."
    email = str(data.get("email", "") or "").strip()
    if email and not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
        return "이메일 형식이 올바르지 않습니다."
    return None


def _safe_image_token(value: Any) -> str:
    token = Path(str(value or "")).name
    if not re.fullmatch(r"[a-f0-9]{32}\.jpg", token):
        return ""
    if not (Path(current_app.config["SCAN_DIR"]) / token).is_file():
        return ""
    return token


def _delete_scan_if_orphaned(token: str) -> None:
    if not token or _database().image_reference_count(token) > 0:
        return
    scan = Path(current_app.config["SCAN_DIR"]) / Path(token).name
    if scan.is_file():
        scan.unlink()


def _export_value(value: Any) -> Any:
    """Prevent spreadsheet programs from interpreting user text as a formula."""
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


@web.get("/")
def index():
    return render_template("index.html")


@web.get("/assets/styles.css")
def frontend_styles():
    return send_file(
        FRONTEND_DIR / "styles.css",
        mimetype="text/css",
        conditional=True,
        max_age=0,
    )


@web.get("/assets/app.js")
def frontend_script():
    return send_file(
        FRONTEND_DIR / "app.js",
        mimetype="text/javascript",
        conditional=True,
        max_age=0,
    )


@web.get("/api/health")
def health():
    return jsonify(
        {
            "ok": True,
            "service": "cardocr",
            "ocr": {
                "engine": "Hybrid OCR",
                "installed": engine.installed(),
                "languages": ["ko", "en"],
                "available_engines": engine.available_engines(),
                **engine.runtime_status(),
            },
            "llm_classifier": gemini_classifier.status(),
            "contacts": _database().count_contacts(),
        }
    )


@web.post("/api/ocr")
def recognize_card():
    started = time.perf_counter()
    upload = request.files.get("image")
    if upload is None or not upload.filename:
        return _error("촬영하거나 선택한 이미지가 필요합니다.")
    extension = Path(upload.filename).suffix.lower()
    if extension and extension not in ALLOWED_EXTENSIONS:
        return _error("JPG, PNG, WEBP, BMP 이미지만 사용할 수 있습니다.")

    try:
        source = decode_image(upload.read())
        decoded_at = time.perf_counter()
        card = detect_and_rectify(source)
        rectified_at = time.perf_counter()
        ocr_image = resize_for_ocr(
            card.image,
            max_long_edge=int(current_app.config["OCR_MAX_LONG_EDGE"]),
        )
        lines = engine.recognize(ocr_image)
        recognized_at = time.perf_counter()
    except InvalidImageError as exc:
        return _error(str(exc))
    except OCRUnavailableError as exc:
        current_app.logger.warning(
            "OCR 실패: %.0fms - %s",
            (time.perf_counter() - started) * 1000,
            exc,
        )
        return _error(str(exc), 503, "OCR_UNAVAILABLE")

    parsed = ParsedCard(raw_text="\n".join(line.text for line in lines))
    parsed, llm_classifier_info = gemini_classifier.enhance(
        lines,
        parsed,
        image_bytes=encode_jpeg(ocr_image, quality=88),
    )
    uncertain_mixed_text = any(
        line.confidence < 0.45
        and bool(re.search(r"[가-힣]", line.text))
        and (
            bool(re.search(r"[A-Za-z]", line.text))
            or max(point[1] for point in line.box) - min(point[1] for point in line.box) < 30
        )
        for line in lines
    )
    if uncertain_mixed_text:
        card.quality_warnings.append(
            "작은 한글 또는 한영 혼합 문장의 인식 신뢰도가 낮습니다. 저장 전에 원문과 비교해 주세요."
        )
    image_token = f"{uuid.uuid4().hex}.jpg"
    (Path(current_app.config["SCAN_DIR"]) / image_token).write_bytes(
        encode_jpeg(card.image, quality=92)
    )
    finished_at = time.perf_counter()
    timings = {
        "decode": round((decoded_at - started) * 1000),
        "rectify": round((rectified_at - decoded_at) * 1000),
        "ocr": round((recognized_at - rectified_at) * 1000),
        "parse_and_save": round((finished_at - recognized_at) * 1000),
        "total": round((finished_at - started) * 1000),
    }
    current_app.logger.info(
        "OCR 완료: total=%dms decode=%dms rectify=%dms ocr=%dms save=%dms, "
        "source=%dx%d input=%dx%d lines=%d",
        timings["total"],
        timings["decode"],
        timings["rectify"],
        timings["ocr"],
        timings["parse_and_save"],
        source.shape[1],
        source.shape[0],
        ocr_image.shape[1],
        ocr_image.shape[0],
        len(lines),
    )

    return jsonify(
        {
            "ok": True,
            "data": {
                "fields": parsed.to_dict(),
                "llm_classifier": llm_classifier_info,
                "ocr_lines": [
                    {"text": line.text, "confidence": line.confidence, "box": line.box}
                    for line in lines
                ],
                "image_token": image_token,
                "preview": as_data_url(card.image),
                "processing_ms": timings,
                "detection": {
                    "card_detected": card.detected,
                    "corners": card.corners,
                    "warnings": card.quality_warnings,
                    **image_metadata(card.image),
                },
            },
        }
    )


@web.post("/api/parse")
def parse_text():
    try:
        data = _payload()
    except ValueError as exc:
        return _error(str(exc))
    raw_text = str(data.get("raw_text", ""))
    if not raw_text.strip():
        return _error("분류할 OCR 텍스트를 입력해 주세요.")
    text_lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    lines = [
        OCRLine(
            text,
            1.0,
            [[0, index * 20], [1000, index * 20], [1000, index * 20 + 18], [0, index * 20 + 18]],
        )
        for index, text in enumerate(text_lines)
    ]
    parsed = ParsedCard(raw_text="\n".join(text_lines))
    parsed, classifier_info = gemini_classifier.enhance(lines, parsed)
    return jsonify(
        {
            "ok": True,
            "data": {
                "fields": parsed.to_dict(),
                "llm_classifier": classifier_info,
            },
        }
    )


@web.get("/api/contacts")
def list_contacts():
    query = request.args.get("q", "")
    contacts = _database().list_contacts(query)
    return jsonify({"ok": True, "data": contacts, "count": len(contacts)})


@web.post("/api/contacts")
def create_contact():
    try:
        data = _payload()
    except ValueError as exc:
        return _error(str(exc))
    problem = _validate_contact(data)
    if problem:
        return _error(problem)
    duplicates = _database().find_duplicates(
        str(data.get("phone", "") or ""),
        str(data.get("email", "") or ""),
        phone2=str(data.get("phone2", "") or ""),
    )
    if duplicates and not bool(data.get("allow_duplicate")):
        return (
            jsonify(
                {
                    "ok": False,
                    "error": {
                        "code": "DUPLICATE_CONTACT",
                        "message": "같은 전화번호 또는 이메일의 고객이 이미 있습니다.",
                        "duplicates": duplicates,
                    },
                }
            ),
            409,
        )
    data["image_token"] = _safe_image_token(data.get("image_token"))
    return jsonify({"ok": True, "data": _database().create_contact(data)}), 201


@web.get("/api/contacts/<int:contact_id>")
def get_contact(contact_id: int):
    contact = _database().get_contact(contact_id)
    if contact is None:
        return _error("고객정보를 찾을 수 없습니다.", 404, "NOT_FOUND")
    return jsonify({"ok": True, "data": contact})


@web.put("/api/contacts/<int:contact_id>")
def update_contact(contact_id: int):
    try:
        data = _payload()
    except ValueError as exc:
        return _error(str(exc))
    problem = _validate_contact(data)
    if problem:
        return _error(problem)
    duplicates = _database().find_duplicates(
        str(data.get("phone", "") or ""),
        str(data.get("email", "") or ""),
        exclude_id=contact_id,
        phone2=str(data.get("phone2", "") or ""),
    )
    if duplicates and not bool(data.get("allow_duplicate")):
        return (
            jsonify(
                {
                    "ok": False,
                    "error": {
                        "code": "DUPLICATE_CONTACT",
                        "message": "같은 전화번호 또는 이메일의 고객이 이미 있습니다.",
                        "duplicates": duplicates,
                    },
                }
            ),
            409,
        )
    existing = _database().get_contact(contact_id)
    if existing is None:
        return _error("고객정보를 찾을 수 없습니다.", 404, "NOT_FOUND")
    image_token = _safe_image_token(data.get("image_token"))
    data["image_token"] = image_token or existing.get("image_token", "")
    contact = _database().update_contact(contact_id, data)
    if existing.get("image_token") != data["image_token"]:
        _delete_scan_if_orphaned(str(existing.get("image_token", "")))
    return jsonify({"ok": True, "data": contact})


@web.delete("/api/contacts/<int:contact_id>")
def delete_contact(contact_id: int):
    existing = _database().get_contact(contact_id)
    if existing is None:
        return _error("고객정보를 찾을 수 없습니다.", 404, "NOT_FOUND")
    if not _database().delete_contact(contact_id):
        return _error("고객정보를 찾을 수 없습니다.", 404, "NOT_FOUND")
    _delete_scan_if_orphaned(str(existing.get("image_token", "")))
    return jsonify({"ok": True})


@web.get("/api/scans/<path:token>")
def get_scan(token: str):
    safe_token = _safe_image_token(token)
    if not safe_token:
        return _error("이미지를 찾을 수 없습니다.", 404, "NOT_FOUND")
    return send_file(Path(current_app.config["SCAN_DIR"]) / safe_token, mimetype="image/jpeg")


@web.get("/api/export")
def export_contacts():
    file_format = request.args.get("format", "csv").lower()
    contacts = _database().list_contacts(request.args.get("q", ""))
    fields = list(EXPORT_HEADERS)

    if file_format == "csv":
        stream = io.StringIO(newline="")
        writer = csv.writer(stream)
        writer.writerow([EXPORT_HEADERS[field] for field in fields])
        writer.writerows(
            [[_export_value(contact.get(field, "")) for field in fields] for contact in contacts]
        )
        payload = io.BytesIO(stream.getvalue().encode("utf-8-sig"))
        return send_file(
            payload,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name="customer_contacts.csv",
        )

    if file_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return _error("Excel 출력을 위해 openpyxl 설치가 필요합니다.", 503, "EXPORT_UNAVAILABLE")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "고객정보"
        sheet.append([EXPORT_HEADERS[field] for field in fields])
        for contact in contacts:
            sheet.append([_export_value(contact.get(field, "")) for field in fields])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center")
        widths = [8, 16, 24, 16, 18, 18, 30, 28, 38, 28, 20, 20]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return send_file(
            payload,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="customer_contacts.xlsx",
        )

    return _error("format은 csv 또는 xlsx여야 합니다.")


@web.errorhandler(413)
def image_too_large(_exception):
    return _error("이미지는 15MB 이하만 업로드할 수 있습니다.", 413, "FILE_TOO_LARGE")
